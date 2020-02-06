#!/usr/bin/python3
from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import datetime
import io
import pyexcel_ods as p
import json
import logging
from string import ascii_uppercase
import traceback
import re
import shelve
from myconfig import *
import openpyxl

''' TODO:
    '''

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive']
alphabet = list(ascii_uppercase)
#CRON_TIME = 5 #the time between one execution and the next
filename_regex = re.compile(r'\d\d-')
file_already_logged = False
#folder_id = "1RvdbykGns22dh7t9q6P0mqINk_Ni0x-T" #the id of the folder where the logfiles are stored

def setup_logger(name, log_file, level=logging.WARNING):

    handler = logging.FileHandler(log_file)
    handler.setFormatter(logging.Formatter('%(asctime)s | %(message)s'))

    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)

    return logger

#computes the difference between current_datetime and the modifiedTime of the file
def minutes_from_last_change(modifiedTime, current_datetime):
    file_time = modifiedTime.split('T')
    file_time_date = file_time[0]
    file_time_hour = file_time[1].split('.')[0]
    file_time_str = file_time_date + " " +file_time_hour
    datetime_object = datetime.datetime.strptime(file_time_str, '%Y-%m-%d %H:%M:%S')
    datetime_difference = current_datetime - datetime_object
    minutes_seconds = divmod(datetime_difference.days * 86400 + datetime_difference.seconds, 60)
    minutes = minutes_seconds[0]
    return minutes

def date_converter(obj):
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.__str__()

def get_revision_index(revisions, current_datetime):
    iter = 0
    while(minutes_from_last_change(revisions[iter]["modifiedTime"], current_datetime) > CRON_TIME):
        iter = iter + 1
    return iter - 1 if iter -1 > 0 else 0 #find the first revision in the CRON_TIME minutes window

class File:
    def __init__(self, service, item, LOG):
        self.service = service #instace of GOOGLE API
        self.item = item #DICT that holds the file information
        self.LOG = LOG #app logger
        self.revision = None #DICT that holds the revision information
        self.file_log = None #file logger
        self.lastModifyingUser = None #the user that last modified the file
        self.revisions = None #list of revisions
        self.modifiedTime = None #the the last modification time

    def set_revision(self, revision):
        self.revision = revision

    #download the ods file from Drive
    def download_file(self, fileid = None, name = None):
        fileid = self.item['id'] if fileid is None else fileid
        name = self.item['name'] if name is None else name
        self.LOG.info('Downloading file %s', name)
        request = self.service.files().get_media(fileId= fileid)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            self.LOG.info("Download file {0}".format(status.progress()*100))

        with open(name, "wb") as out:
            out.write(fh.getvalue())

    #download the revision from drive
    def download_revision(self, revision_id, name):
        if name in os.listdir():
            return
        self.LOG.info('Downloading revision %s', name)
        request = self.service.revisions().get_media(revisionId=revision_id, fileId=self.item["id"])
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            self.LOG.info("Download revision {0}".format(status.progress()*100))


        with open(name, "wb") as out:
            out.write(fh.getvalue())

    #setup the logger" for the current file
    def setup_logger(self, name = None, log_file=None, level=logging.WARNING):
        log_file = self.item["name"]+".log" if log_file is None else log_file
        handler = logging.FileHandler(log_file)
        handler.setFormatter(logging.Formatter('%(asctime)s | %(message)s'))

        name = self.item["name"] if name is None else name
        logger = logging.getLogger(name)
        logger.setLevel(level)
        logger.addHandler(handler)

        self.file_log = logger

    def __calculate_ods_coordinate(self, row, column):
        return (alphabet[column] + str(row))

    #get the diffrence between each row of the file and the corresponding row in the revision
    def __get_difference_rows(self, row_current, row_modified, index):
        min_range = min(len(row_current), len(row_modified))
        for i in range(0, min_range):
            if (row_current[i] != row_modified[i]):
                text_current = row_current[i]
                if(text_current == ""):
                    text_current = "\"\""
                text_modified = row_modified[i]
                if(text_modified == ""):
                    text_modified = "\"\""
                cell_coordinates = self.__calculate_ods_coordinate(index+1, i)
                # if "displayName" in self.item["lastModifyingUser"].keys():
                #     self.file_log.info("{0} changed from {1} to {2} by {3}". format(cell_coordinates, text_modified, text_current, self.item["lastModifyingUser"]["displayName"]))
                # else:
                #     if self.lastModifyingUser is None or self.lastModifyingUser == "not found":
                #         results = self.service.files().get(fileId=self.item["id"], fields="lastModifyingUser").execute()
                #         if "displayName" in results['lastModifyingUser'].keys():
                #             self.lastModifyingUser = results["lastModifyingUser"]["displayName"]
                #         else:
                #             self.LOG.info("user not found")
                #             self.lastModifyingUser = "not found"
                #     self.file_log.info("{0} changed from {1} to {2} by {3}". format(cell_coordinates, text_modified, text_current, self.lastModifyingUser))
                self.file_log.info("{0} changed from {1} to {2} by {3} at {4}". format(cell_coordinates, text_modified, text_current, self.lastModifyingUser, self.modifiedTime))

        if(len(row_current) > len(row_modified)):
            i = min_range
            while(i != len(row_current)):
                if(row_current[i]):
                    cell_coordinates = self.__calculate_ods_coordinate(index+1, i)
                    # if "displayName" in self.item["lastModifyingUser"].keys():
                    #     self.file_log.info("{0} changed from \"\" to {1} by {2}".format(cell_coordinates, row_current[i], self.item["lastModifyingUser"]["displayName"]))
                    # else:
                    #     if self.lastModifyingUser is None or self.lastModifyingUser == "not found":
                    #         results = self.service.files().get(fileId=self.item["id"], fields="lastModifyingUser").execute()
                    #         if "displayName" in results['lastModifyingUser'].keys():
                    #             self.lastModifyingUser = results["lastModifyingUser"]["displayName"]
                    #         else:
                    #             self.LOG.info("user not found")
                    #             self.lastModifyingUser = "not found"
                    #     self.file_log.info("{0} changed from \"\" to {1} by {2}".format(cell_coordinates, row_current[i], self.lastModifyingUser))
                    self.file_log.info("{0} changed from \"\" to {1} by {2} at {3}".format(cell_coordinates, row_current[i], self.lastModifyingUser, self.modifiedTime))
                i += 1
        elif(len(row_current) < len(row_modified)):
            i = min_range
            while(i != len(row_modified)):
                if(row_modified[i]):
                    cell_coordinates = self.__calculate_ods_coordinate(index+1, i)
                    # if "displayName" in self.item["lastModifyingUser"].keys():
                    #     self.file_log.info("{0} changed from {1} to \"\" by {2}".format(cell_coordinates, row_modified[i], self.item["lastModifyingUser"]["displayName"]))
                    # else:
                    #     if self.lastModifyingUser is None or self.lastModifyingUser == "not found":
                    #         results = self.service.files().get(fileId=self.item["id"], fields="lastModifyingUser").execute()
                    #         if "displayName" in results['lastModifyingUser'].keys():
                    #             self.lastModifyingUser = results["lastModifyingUser"]["displayName"]
                    #         else:
                    #             self.LOG.info("user not found")
                    #             self.lastModifyingUser = "not found"
                    #     self.file_log.info("{0} changed from {1} to \"\" by {2}".format(cell_coordinates, row_modified[i], self.lastModifyingUser))
                    self.file_log.info("{0} changed from {1} to \"\" by {2} at {3}".format(cell_coordinates, row_modified[i], self.lastModifyingUser, self.modifiedTime))
                i += 1

    #read the ods file as a dict and call __get_difference_rows on every row
    def get_difference(self, name1, name2):
        self.file_log.info("<---------BEGIN LOG--------->")
        data_current = p.get_data(name2)
        json_string_current = json.dumps(data_current, default=date_converter)
        json_dict_current = json.loads(json_string_current)

        data_modified = p.get_data(name1)
        json_string_modified = json.dumps(data_modified, default=date_converter)
        json_dict_modified = json.loads(json_string_modified)

        for key in json_dict_current.keys():
            sheet_current = json_dict_current[key]
            sheet_modified = json_dict_modified[key]
            if key != "modifiche":
                for i in range(0, len(sheet_current)):
                    row_current = sheet_current[i]
                    if i < len(sheet_modified):
                        row_modified = sheet_modified[i]
                    else:
                        row_modified = "";
                    self.__get_difference_rows(row_current, row_modified, i)
        self.file_log.info("<---------END LOG--------->")

    def file_created(self):
        metadata = self.service.revisions().get(fileId=self.item["id"], revisionId=self.revisions[-1]["id"], fields="lastModifyingUser, modifiedTime").execute()
        modifiedTime = metadata["modifiedTime"]
        file_time = modifiedTime.split('T')
        file_time_date = file_time[0]
        file_time_hour = file_time[1].split('.')[0]
        time_list = file_time_hour.split(':')
        correct_hour = int(time_list[0]) + 1
        file_time_hour = str(correct_hour) + ":" + time_list[1] + ":" +time_list[2]
        file_time_str = file_time_date + " " +file_time_hour
        self.file_log.info("{0} has created the file at {1}".format(metadata["lastModifyingUser"]["displayName"], file_time_str))

    def set_lastModifyingUser(self, username):
        self.lastModifyingUser = username

    def set_modifiedTime(self,modifiedTime):
        file_time = modifiedTime.split('T')
        file_time_date = file_time[0]
        file_time_hour = file_time[1].split('.')[0]
        time_list = file_time_hour.split(':')
        correct_hour = (int(time_list[0]) + 1) % 24
        file_time_hour = str(correct_hour) + ":" + time_list[1] + ":" +time_list[2]
        file_time_str = file_time_date + " " +file_time_hour
        self.modifiedTime = file_time_str

    def compute_revisions(self, last_revision_index):
        for i in range(last_revision_index, len(self.revisions)-1):
            name1 = "revision" + str(i) + "_" + self.item["name"]
            name2 = "revision" + str(i+1) + "_" + self.item["name"]
            metadata = self.service.revisions().get(fileId=self.item["id"], revisionId=self.revisions[i+1]["id"], fields="lastModifyingUser, modifiedTime").execute()
            lastModifyingUser = metadata["lastModifyingUser"]["displayName"]
            self.set_lastModifyingUser(lastModifyingUser)
            self.set_modifiedTime(metadata["modifiedTime"])
            self.download_revision(self.revisions[i]['id'], name1)
            self.download_revision(self.revisions[i+1]['id'], name2)
            self.get_difference(name1, name2)
            if(os.path.exists(name1)):
                os.remove(name1)
            else:
                self.LOG.info("file {0} does not exist".format(name1))

    def set_revisions(self, revisions):
        self.revisions = revisions


class Excel_File(File):
    def __init__(self, service, item, LOG):
        super(Excel_File, self).__init__(service, item, LOG)
        self.wb_current = None
        self.wb_modified = None

    def __get_difference_rows(self, row_current, row_modified):
        row_current = list(row_current)
        row_modified = list(row_modified)
        min_range = min(len(row_current), len(row_modified))
        for i in range(0, min_range):
            current_cell = row_current[i]
            modified_cell = row_modified[i]
            if (current_cell.value != modified_cell.value):
                text_current = current_cell.value
                if(text_current == ""):
                    text_current = "\"\""
                text_modified = modified_cell.value
                if(text_modified == ""):
                    text_modified = "\"\""

                self.file_log.info("{0} changed from {1} to {2} by {3} at {4}".format(current_cell.coordinate, text_modified, text_current, self.lastModifyingUser, self.modifiedTime))
        if(len(row_current) > len(row_modified)):
            i = min_range
            while(i != len(row_current)):
                current_cell = row_current[i]
                if(current_cell.value):
                    self.file_log.info("{0} changed from \"\" to {1} by {2} at {3}".format(current_cell.coordinate, current_cell.value, self.lastModifyingUser, self.modifiedTime))
                i +=1
        elif(len(row_current) < len(row_modified)):
            i = min_range
            while(i != len(row_modified)):
                modified_cell = row_modified[i]
                if(modified_cell.value):
                    self.file_log.info("{0} changed from {1} to \"\" by {2} at {3}".format(modified_cell.coordinate, modified_cell.value, self.lastModifyingUser, self.modifiedTime))
                i +=1


    def get_difference(self, name1, name2):
        self.file_log.info("<---------BEGIN LOG--------->")
        self.wb_current = openpyxl.load_workbook(name2)
        self.wb_modified = openpyxl.load_workbook(name1)
        sheets = self.wb_current.get_sheet_names()
        for name in sheets:
            if (name != "modifiche"):

                sheet_current = self.wb_current.get_sheet_by_name(name)
                sheet_modified = self.wb_modified.get_sheet_by_name(name)
                for count, row_current in enumerate(sheet_current.iter_rows()):
                    row_modified = sheet_modified[count + 1]
                    self.__get_difference_rows(row_current, row_modified)



        self.file_log.info("<---------END LOG--------->")




def remove_file(filename, LOG):
    if(os.path.exists(filename)):
        os.remove(filename)
    else:
        LOG.error("file {0} does not exist".format(filename))

def main():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('drive', 'v3', credentials=creds)

    LOG = setup_logger("app.log", "app.log", logging.INFO)

    page_token = None

    try:
        shelfFile = shelve.open('log_list')
    except err as e:
        LOG.error('error in opening the shelve file')

    global file_already_logged

    #make a request until there are no more files to process
    while True:
        # Call the Drive v3 API
        response = service.files().list(
            q = "(mimeType='application/vnd.oasis.opendocument.spreadsheet' or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') and trashed=false", pageSize=1000, fields="nextPageToken, files(id,name,modifiedTime,lastModifyingUser,mimeType)", pageToken=page_token).execute()
        items = response.get('files', [])

        current_datetime = datetime.datetime.utcnow()

        if not items:
            LOG.info("No files found using the provided query")
        else:
            for item in items:
                re_results = filename_regex.search(item["name"])
                if re_results is not None:
                    minutes = minutes_from_last_change(item["modifiedTime"], current_datetime)
                    if(minutes < CRON_TIME):
                        if (item["mimeType"] == "application/vnd.oasis.opendocument.spreadsheet"):
                            my_file = File(service, item, LOG)
                        else:
                            my_file = Excel_File(service, item, LOG)
                        #my_file.download_file()

                        results = service.revisions().list(fileId=item["id"]).execute()
                        revisions = results.get("revisions", [])
                        my_file.setup_logger(level=logging.INFO)
                        my_file.set_revisions(revisions)

                        if(len(revisions) > 1):
                            log_name = item['name'] + '.log'
                            log_keys = shelfFile.keys()
                            log_id = None
                            if log_name in log_keys:
                                log_id = shelfFile[log_name]
                                my_file.download_file(log_id, log_name)
                                file_already_logged = True

                            revision_index = get_revision_index(revisions, current_datetime)
                            #my_file.set_revision(revisions[revision_index])
                            #my_file.download_revision()
                            try:
                                my_file.compute_revisions(revision_index)
                                log_metadata = {'name' : log_name, 'parents' : [folder_id]}
                                media = MediaFileUpload(log_name, mimetype='text/plain', resumable=True)
                                if file_already_logged:
                                    log_metadata = {'name' : log_name} #the parents field in the metadata is not writable using the update request, use addParent,removeParents instead
                                    file = service.files().update(fileId=log_id, body=log_metadata, media_body=media, fields='id').execute()
                                    LOG.info('{} log updated successfully'.format(log_name))
                                else:
                                    file = service.files().create(body=log_metadata, media_body=media, fields='id').execute()
                                    shelfFile[log_name] = file.get('id')
                                    LOG.info('{} log created successfully'.format(log_name))
                                    permissions_metadata = {"type": "anyone", "role": "reader"}
                                    permissions = service.permissions().create(fileId=file.get('id'), body=permissions_metadata).execute()
                                remove_file(log_name, LOG)
                            except KeyError:
                                LOG.info("error in reading the files content")
                                traceback.print_exc()
                                remove_file(item["name"]+".log", LOG)
                            except Exception as e:
                                LOG.info("error : {0}".format(str(e)))
                                traceback.print_exc()

                                remove_file(item["name"]+".log", LOG)
                            #remove_file("revision_" + item["name"], LOG)
                        else:
                            my_file.file_created()
                            logname = item['name'] + '.log'
                            log_metadata = {'name' : logname, 'parents' : [folder_id]}
                            media = MediaFileUpload(logname, mimetype='text/plain', resumable=True)
                            file = service.files().create(body=log_metadata, media_body=media, fields='id').execute()
                            permissions_metadata = {"type": "anyone", "role": "reader"}
                            permissions = service.permissions().create(fileId=file.get('id'), body=permissions_metadata).execute()
                            shelfFile[logname] = file.get('id')
                            LOG.info("{} log created and uploaded".format(logname))
                            remove_file(logname, LOG)
                        name_last_revision = "revision" + str(len(revisions)-1) + "_" + item["name"]
                        remove_file(name_last_revision, LOG)
                        file_already_logged = False

                        del my_file

        page_token = response.get('nextPageToken', None)
        if page_token is None:
            break


if __name__ == '__main__':
    main()
